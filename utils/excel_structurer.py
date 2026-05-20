from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import math

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


EXCEL_FILE_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".xlsb"}

_PREVIEW_LINE_LIMIT = 60
_HEADER_SCAN_LIMIT = 10
_SUMMARY_HEADER_LIMIT = 8
_TOP_LEVEL_SHARED_FIELD_LIMIT = 40


def is_excel_file(path_or_name: str) -> bool:
    return Path(str(path_or_name or "")).suffix.lower() in EXCEL_FILE_EXTENSIONS


def _is_nan(value: Any) -> bool:
    try:
        return bool(pd.isna(value))
    except Exception:
        return False


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if _is_nan(value):
        return True
    return not str(value).strip()


def _stringify_value(value: Any) -> str:
    if _is_blank(value):
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return format(value, "g")
    return str(value).strip()


def _normalize_header_name(value: Any, index: int, seen: dict[str, int]) -> str:
    base = _stringify_value(value) or f"column_{index + 1}"
    count = seen.get(base, 0) + 1
    seen[base] = count
    return base if count == 1 else f"{base}_{count}"


def _normalize_metadata_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    normalized = []
    previous_was_sep = False
    for char in text:
        if char.isalnum():
            normalized.append(char)
            previous_was_sep = False
            continue
        if not previous_was_sep:
            normalized.append("_")
            previous_was_sep = True
    return "".join(normalized).strip("_")


def _row_non_empty_cells(row: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [cell for cell in row if str(cell.get("value") or "").strip() or str(cell.get("formula") or "").strip()]


def _looks_like_number(text: str) -> bool:
    candidate = str(text or "").strip().replace(",", "")
    if not candidate:
        return False
    try:
        float(candidate)
        return True
    except Exception:
        return False


def _detect_header_row(rows: list[dict[str, Any]]) -> int | None:
    for idx, row in enumerate(rows[:_HEADER_SCAN_LIMIT]):
        cells = row.get("cells") or []
        values = [str(cell.get("value") or "").strip() for cell in cells if str(cell.get("value") or "").strip()]
        if len(values) < 2:
            continue

        distinct_ratio = len({value.lower() for value in values}) / max(len(values), 1)
        textish_ratio = sum(0 if _looks_like_number(value) else 1 for value in values) / max(len(values), 1)
        if distinct_ratio < 0.6 or textish_ratio < 0.5:
            continue

        populated_following_rows = 0
        for next_row in rows[idx + 1: idx + 4]:
            next_values = [str(cell.get("value") or "").strip() for cell in next_row.get("cells") or [] if str(cell.get("value") or "").strip()]
            if len(next_values) >= 2:
                populated_following_rows += 1
        if populated_following_rows >= 1:
            return idx
    return None


def _looks_like_key_value_sheet(rows: list[dict[str, Any]]) -> bool:
    key_value_rows = 0
    for row in rows[: min(len(rows), 25)]:
        cells = _row_non_empty_cells(row.get("cells") or [])
        if len(cells) == 2:
            key_value_rows += 1
    return key_value_rows >= 3


def _build_tabular_region(sheet_name: str, rows: list[dict[str, Any]], header_index: int) -> dict[str, Any]:
    header_row = rows[header_index]
    seen: dict[str, int] = {}
    headers = [
        _normalize_header_name(cell.get("value"), idx, seen)
        for idx, cell in enumerate(header_row.get("cells") or [])
    ]

    region_rows = []
    for row in rows[header_index + 1:]:
        cells = row.get("cells") or []
        values_map: dict[str, str] = {}
        formulas_map: dict[str, str] = {}

        for idx, cell in enumerate(cells):
            if idx >= len(headers):
                break
            header = headers[idx]
            value = str(cell.get("value") or "").strip()
            formula = str(cell.get("formula") or "").strip()
            if value:
                values_map[header] = value
            if formula:
                formulas_map[header] = formula

        if not values_map and not formulas_map:
            continue

        payload = {"_row": row.get("row")}
        payload.update(values_map)
        if formulas_map:
            payload["_formulas"] = formulas_map
        region_rows.append(payload)

    return {
        "region_id": f"{sheet_name}:table:1",
        "type": "tabular",
        "header_row": header_row.get("row"),
        "columns": headers,
        "rows": region_rows,
    }


def _extract_tabular_metadata(region: dict[str, Any]) -> dict[str, Any]:
    rows = region.get("rows") or []
    columns = region.get("columns") or []
    shared_fields: dict[str, str] = {}

    for column in columns:
        seen_values = []
        seen_lookup = set()
        for row in rows:
            value = str(row.get(column) or "").strip()
            if not value:
                continue
            key = value.lower()
            if key in seen_lookup:
                continue
            seen_lookup.add(key)
            seen_values.append(value)
            if len(seen_values) > 1:
                break
        if len(seen_values) == 1:
            shared_fields[column] = seen_values[0]

    top_level_fields: dict[str, Any] = {}
    for idx, (column, value) in enumerate(shared_fields.items()):
        if idx >= _TOP_LEVEL_SHARED_FIELD_LIMIT:
            break
        key = _normalize_metadata_key(column)
        if key and key not in top_level_fields:
            top_level_fields[key] = value

    top_level_fields["items"] = rows
    top_level_fields["row_count"] = len(rows)
    top_level_fields["column_headers"] = [str(column or "").strip() for column in columns if str(column or "").strip()]
    top_level_fields["shared_fields"] = shared_fields
    return top_level_fields


def _build_key_value_region(sheet_name: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    entries = []
    for row in rows:
        cells = _row_non_empty_cells(row.get("cells") or [])
        if len(cells) < 2:
            continue
        key = str(cells[0].get("value") or "").strip()
        value = str(cells[1].get("value") or "").strip()
        if not key or not value:
            continue
        payload = {
            "_row": row.get("row"),
            "key": key,
            "value": value,
        }
        formula = str(cells[1].get("formula") or "").strip()
        if formula:
            payload["formula"] = formula
        entries.append(payload)

    return {
        "region_id": f"{sheet_name}:key_value:1",
        "type": "key_value",
        "rows": entries,
    }


def _build_raw_region(sheet_name: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    raw_rows = []
    for row in rows:
        cells = _row_non_empty_cells(row.get("cells") or [])
        if not cells:
            continue
        raw_rows.append(
            {
                "_row": row.get("row"),
                "cells": [
                    {
                        "column": cell.get("column"),
                        "value": cell.get("value"),
                        **({"formula": cell.get("formula")} if cell.get("formula") else {}),
                    }
                    for cell in cells
                ],
            }
        )

    return {
        "region_id": f"{sheet_name}:raw:1",
        "type": "raw_grid",
        "rows": raw_rows,
    }


def _iter_openpyxl_sheet_rows(file_path: str, keep_vba: bool = False) -> list[dict[str, Any]]:
    wb_values = load_workbook(file_path, data_only=True, read_only=True, keep_vba=keep_vba)
    wb_formulas = load_workbook(file_path, data_only=False, read_only=True, keep_vba=keep_vba)

    try:
        workbook_sheets = []
        for values_ws, formulas_ws in zip(wb_values.worksheets, wb_formulas.worksheets):
            rows = []
            max_col = max(values_ws.max_column or 0, formulas_ws.max_column or 0)
            for row_index, (value_row, formula_row) in enumerate(
                zip(
                    values_ws.iter_rows(min_row=1, max_col=max_col),
                    formulas_ws.iter_rows(min_row=1, max_col=max_col),
                ),
                start=1,
            ):
                row_cells = []
                for col_index, (value_cell, formula_cell) in enumerate(zip(value_row, formula_row), start=1):
                    value = _stringify_value(value_cell.value)
                    formula = ""
                    raw_formula = formula_cell.value
                    if isinstance(raw_formula, str) and raw_formula.startswith("="):
                        formula = raw_formula.strip()
                    if not value and not formula:
                        row_cells.append(
                            {
                                "column": get_column_letter(col_index),
                                "value": "",
                                "formula": "",
                            }
                        )
                        continue
                    row_cells.append(
                        {
                            "column": get_column_letter(col_index),
                            "value": value,
                            "formula": formula,
                        }
                    )
                rows.append({"row": row_index, "cells": row_cells})

            workbook_sheets.append(
                {
                    "name": values_ws.title,
                    "visibility": getattr(values_ws, "sheet_state", "visible") or "visible",
                    "max_row": max(values_ws.max_row or 0, formulas_ws.max_row or 0),
                    "max_col": max_col,
                    "merged_ranges": [],
                    "rows": rows,
                }
            )
        return workbook_sheets
    finally:
        wb_values.close()
        wb_formulas.close()


def _iter_pandas_sheet_rows(file_path: str) -> list[dict[str, Any]]:
    workbook_sheets = []
    excel_file = pd.ExcelFile(file_path)
    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, dtype=object)
        rows = []
        max_row = int(df.shape[0])
        max_col = int(df.shape[1])
        for row_index in range(max_row):
            row_cells = []
            for col_index in range(max_col):
                row_cells.append(
                    {
                        "column": get_column_letter(col_index + 1),
                        "value": _stringify_value(df.iat[row_index, col_index]),
                        "formula": "",
                    }
                )
            rows.append({"row": row_index + 1, "cells": row_cells})

        workbook_sheets.append(
            {
                "name": sheet_name,
                "visibility": "visible",
                "max_row": max_row,
                "max_col": max_col,
                "merged_ranges": [],
                "rows": rows,
            }
        )
    return workbook_sheets


def _sheet_to_canonical(sheet_payload: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    rows = sheet_payload.get("rows") or []
    non_empty_rows = []
    non_empty_cell_count = 0

    for row in rows:
        raw_cells = row.get("cells") or []
        cells = _row_non_empty_cells(raw_cells)
        if not cells:
            continue
        non_empty_rows.append({"row": row.get("row"), "cells": raw_cells})
        non_empty_cell_count += len(cells)

    header_index = _detect_header_row(non_empty_rows)
    if header_index is not None:
        kind = "tabular"
        region = _build_tabular_region(sheet_payload["name"], non_empty_rows, header_index)
    elif _looks_like_key_value_sheet(non_empty_rows):
        kind = "key_value"
        region = _build_key_value_region(sheet_payload["name"], non_empty_rows)
    else:
        kind = "raw_grid"
        region = _build_raw_region(sheet_payload["name"], non_empty_rows)

    retrieval_lines = _build_region_retrieval_lines(sheet_payload["name"], kind, region)

    sheet = {
        "name": sheet_payload["name"],
        "visibility": sheet_payload.get("visibility", "visible"),
        "dimensions": {
            "max_row": int(sheet_payload.get("max_row") or 0),
            "max_col": int(sheet_payload.get("max_col") or 0),
        },
        "merged_ranges": list(sheet_payload.get("merged_ranges") or []),
        "kind": kind,
        "non_empty_row_count": len(non_empty_rows),
        "non_empty_cell_count": non_empty_cell_count,
        "regions": [region],
    }
    return sheet, retrieval_lines


def _build_region_retrieval_lines(sheet_name: str, kind: str, region: dict[str, Any]) -> list[str]:
    lines = []
    if kind == "tabular":
        lines.append(f"Sheet: {sheet_name} | Type: tabular")
        for row in region.get("rows") or []:
            parts = [f"Sheet: {sheet_name}", f"Row: {row.get('_row')}"]
            for column in region.get("columns") or []:
                value = str(row.get(column) or "").strip()
                if value:
                    parts.append(f"{column}: {value}")
            if isinstance(row.get("_formulas"), dict):
                for column, formula in row["_formulas"].items():
                    if formula and not row.get(column):
                        parts.append(f"{column} formula: {formula}")
            if len(parts) > 2:
                lines.append(" | ".join(parts))
        return lines

    if kind == "key_value":
        lines.append(f"Sheet: {sheet_name} | Type: key_value")
        for row in region.get("rows") or []:
            parts = [
                f"Sheet: {sheet_name}",
                f"Row: {row.get('_row')}",
                f"{row.get('key')}: {row.get('value')}",
            ]
            if row.get("formula"):
                parts.append(f"Formula: {row.get('formula')}")
            lines.append(" | ".join(parts))
        return lines

    lines.append(f"Sheet: {sheet_name} | Type: raw_grid")
    for row in region.get("rows") or []:
        parts = [f"Sheet: {sheet_name}", f"Row: {row.get('_row')}"]
        for cell in row.get("cells") or []:
            value = str(cell.get("value") or "").strip()
            if value:
                parts.append(f"{cell.get('column')}: {value}")
            formula = str(cell.get("formula") or "").strip()
            if formula and not value:
                parts.append(f"{cell.get('column')} formula: {formula}")
        if len(parts) > 2:
            lines.append(" | ".join(parts))
    return lines


def _build_sheet_summary(sheet: dict[str, Any]) -> dict[str, Any]:
    columns: list[str] = []
    regions = sheet.get("regions") or []
    if regions:
        first_region = regions[0]
        if isinstance(first_region.get("columns"), list):
            columns = [str(item or "").strip() for item in first_region.get("columns") or [] if str(item or "").strip()]

    return {
        "name": sheet.get("name"),
        "kind": sheet.get("kind"),
        "row_count": int(sheet.get("non_empty_row_count") or 0),
        "cell_count": int(sheet.get("non_empty_cell_count") or 0),
        "column_count": int(sheet.get("dimensions", {}).get("max_col") or 0),
        "headers": columns[:_SUMMARY_HEADER_LIMIT],
    }


def parse_excel_to_structured_content(file_path: str, display_name: str = "") -> dict[str, Any]:
    suffix = Path(file_path).suffix.lower()
    if suffix not in EXCEL_FILE_EXTENSIONS:
        raise ValueError(f"Unsupported Excel file type: {suffix}")

    if suffix in {".xlsx", ".xlsm"}:
        raw_sheets = _iter_openpyxl_sheet_rows(file_path, keep_vba=(suffix == ".xlsm"))
        parser_name = "openpyxl"
    else:
        raw_sheets = _iter_pandas_sheet_rows(file_path)
        parser_name = "pandas"

    workbook_sheets = []
    retrieval_lines = []
    total_non_empty_rows = 0
    total_non_empty_cells = 0
    workbook_items: list[dict[str, Any]] = []
    workbook_shared_fields: dict[str, Any] = {}

    for raw_sheet in raw_sheets:
        sheet, sheet_lines = _sheet_to_canonical(raw_sheet)
        workbook_sheets.append(sheet)
        retrieval_lines.extend(sheet_lines)
        total_non_empty_rows += int(sheet.get("non_empty_row_count") or 0)
        total_non_empty_cells += int(sheet.get("non_empty_cell_count") or 0)

        if sheet.get("kind") == "tabular":
            regions = sheet.get("regions") or []
            if regions:
                extracted = _extract_tabular_metadata(regions[0])
                for item in extracted.get("items") or []:
                    workbook_items.append(
                        {
                            "sheet_name": sheet.get("name"),
                            **item,
                        }
                    )
                shared_fields = extracted.get("shared_fields") if isinstance(extracted.get("shared_fields"), dict) else {}
                for key, value in extracted.items():
                    if key in {"items", "shared_fields"}:
                        continue
                    if key not in workbook_shared_fields:
                        workbook_shared_fields[key] = value
                if shared_fields:
                    workbook_shared_fields.setdefault("sheet_shared_fields", {})
                    workbook_shared_fields["sheet_shared_fields"][sheet.get("name")] = shared_fields

    retrieval_text = "\n".join(line for line in retrieval_lines if str(line).strip()).strip()
    preview_text = "\n".join(retrieval_lines[:_PREVIEW_LINE_LIMIT]).strip()

    sheet_summaries = [_build_sheet_summary(sheet) for sheet in workbook_sheets]
    workbook_payload = {
        "source_type": "excel",
        "parser": parser_name,
        "file_name": str(display_name or Path(file_path).name).strip() or Path(file_path).name,
        "sheet_count": len(workbook_sheets),
        "sheet_names": [sheet.get("name") for sheet in workbook_sheets],
        "total_non_empty_rows": total_non_empty_rows,
        "total_non_empty_cells": total_non_empty_cells,
        "sheets": workbook_sheets,
    }

    return {
        "source_type": "excel",
        "document_type": "excel_workbook",
        "high_level_metadata": {
            "source_type": "excel",
            "file_kind": "workbook",
            "sheet_count": len(workbook_sheets),
            "sheet_names": [sheet.get("name") for sheet in workbook_sheets],
            "total_non_empty_rows": total_non_empty_rows,
            "total_non_empty_cells": total_non_empty_cells,
            "sheet_summaries": sheet_summaries,
            **workbook_shared_fields,
            "items": workbook_items,
        },
        "workbook": workbook_payload,
        "retrieval_text": retrieval_text,
        "preview_text": preview_text or retrieval_text,
    }
