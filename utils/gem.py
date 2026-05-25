"""
GEM (Goods and Services Exchange Model) PDF Processing Module
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from typing import Any, Dict, List, Optional, Tuple, TypedDict
from uuid import uuid4
import fitz
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from langchain_core.prompts import ChatPromptTemplate
from langchain_google_genai import ChatGoogleGenerativeAI
from langgraph.checkpoint.memory import MemorySaver
from langgraph.graph import StateGraph
from rapidfuzz import fuzz
from utils.logger import get_logger


logger = get_logger('gem')
load_dotenv()


# ---------------------------------------------------------------------------
# State Definition for LangGraph
# ---------------------------------------------------------------------------

class ChunkProcessingState(TypedDict):
    all_chunks: List[str]
    current_chunk_index: int
    accumulated_results: Dict[str, Dict]
    master_schema: Dict[str, Any]
    chunk_results: List[Dict[str, Dict]]
    is_complete: bool


# ---------------------------------------------------------------------------
# Text Chunking
# ---------------------------------------------------------------------------

def split_text_into_chunks(text: str, chunk_size: int = 180000, overlap: int = 500) -> List[str]:
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        chunks.append(chunk.strip())
        start = end - overlap if end < len(text) else end
    logger.info(f"[GEM] Text split into {len(chunks)} chunks (size: {chunk_size}, overlap: {overlap})")
    return chunks


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_sheet(wb, target_name: str):
    target = target_name.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]
    return None


# ---------------------------------------------------------------------------
# OCR helper for scanned PDFs
# ---------------------------------------------------------------------------

def _ocr_pdf_bytes(pdf_bytes: bytes, url: str) -> str:
    """
    Render each PDF page to image using PyMuPDF, then OCR with shared PaddleOCR helper.
    Returns extracted text, or empty string if OCR fails.
    """
    def _extract_page_text(result: Any) -> str:
        if not result:
            return ""
        if isinstance(result, dict):
            raw_text = str(result.get("raw_text") or "").strip()
            if raw_text:
                return raw_text
            rec_texts = result.get("rec_texts")
            if isinstance(rec_texts, list):
                return "\n".join(str(item).strip() for item in rec_texts if str(item).strip()).strip()
            text = str(result.get("text") or "").strip()
            if text:
                return text
        if hasattr(result, "get"):
            try:
                raw_text = str(result.get("raw_text") or "").strip()
                if raw_text:
                    return raw_text
            except Exception:
                pass
        if isinstance(result, list):
            lines = []
            for item in result:
                if isinstance(item, dict):
                    text = str(item.get("text") or item.get("raw_text") or "").strip()
                    if text:
                        lines.append(text)
                        continue
                    rec_texts = item.get("rec_texts")
                    if isinstance(rec_texts, list):
                        lines.extend(str(text).strip() for text in rec_texts if str(text).strip())
                        continue
                if isinstance(item, (list, tuple)) and len(item) >= 2:
                    candidate = item[1]
                    if isinstance(candidate, (list, tuple)) and candidate:
                        text = str(candidate[0] or "").strip()
                        if text:
                            lines.append(text)
                            continue
                text = str(item or "").strip()
                if text and text not in {"[]", "{}"}:
                    lines.append(text)
            return "\n".join(lines).strip()
        return ""

    try:
        from PIL import Image
        from ocr.paddle_ocr_engine import paddle_ocr_extract
    except Exception as exc:
        logger.warning(f"[GEM][OCR] Shared PaddleOCR engine unavailable: {exc}")
        return ""

    text_parts = []
    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            page_count = doc.page_count
            logger.info(f"[GEM][OCR] Starting PyMuPDF text extraction on {page_count} pages | URL: {url}")

            for page_no in range(page_count):
                page = doc.load_page(page_no)
                page_parts = []
                try:
                    tabs = page.find_tables()
                    for tab in tabs:
                        df = tab.to_pandas()
                        if not df.empty:
                            page_parts.append(df.to_csv(index=False, sep="|"))
                except Exception:
                    pass
                plain = page.get_text("text").strip()
                if plain:
                    page_parts.append(plain)
                if page_parts:
                    text_parts.append("\n".join(page_parts))
                    logger.info(
                        f"[GEM][OCR] Page {page_no + 1}/{page_count} - "
                        f"{len(page_parts)} content parts extracted (PyMuPDF)"
                    )
                else:
                    logger.warning(
                        f"[GEM][OCR] Page {page_no + 1}/{page_count} - "
                        f"no text detected via PyMuPDF"
                    )

        pymupdf_text = "\n\n".join(text_parts).strip()
        logger.info(
            f"[GEM][OCR] PyMuPDF complete - {len(pymupdf_text)} chars "
            f"from {page_count} pages | URL: {url}"
        )

        if pymupdf_text and len(pymupdf_text) > 0:
            return pymupdf_text

        logger.warning(
            f"[GEM][OCR] PyMuPDF returned 0 chars — falling back to PaddleOCR per-page extraction | URL: {url}"
        )
        try:
            from PIL import Image
            from ocr.paddle_ocr_engine import paddle_ocr_extract
        except Exception as exc:
            logger.warning(f"[GEM][OCR] Shared PaddleOCR engine unavailable: {exc}")
            return ""

        ocr_parts = []
        with fitz.open(stream=pdf_bytes, filetype="pdf") as ocr_doc:
            for page_no in range(ocr_doc.page_count):
                page = ocr_doc.load_page(page_no)
                mat = fitz.Matrix(2.0, 2.0)
                pix = page.get_pixmap(matrix=mat)
                mode = "RGBA" if pix.n == 4 else "RGB"
                img = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
                if mode == "RGBA":
                    img = img.convert("RGB")
                result = paddle_ocr_extract(img)
                page_text = _extract_page_text(result)
                if page_text:
                    ocr_parts.append(page_text)
                    logger.info(
                        f"[GEM][OCR] Page {page_no + 1}/{page_count} - "
                        f"{len(page_text)} chars extracted (PaddleOCR)"
                    )
                else:
                    logger.warning(
                        f"[GEM][OCR] Page {page_no + 1}/{page_count} - "
                        f"no text detected via PaddleOCR | result_type={type(result).__name__}"
                    )

        ocr_result = "\n\n".join(ocr_parts).strip()
        logger.info(
            f"[GEM][OCR] PaddleOCR fallback complete - {len(ocr_result)} chars "
            f"from {page_count} pages | URL: {url}"
        )
        return ocr_result

    except Exception as e:
        logger.error(f"[GEM][OCR] Text extraction failed: {e} | URL: {url}")
        return ""


# ---------------------------------------------------------------------------
# HTML scraping find PDF links inside HTML pages
# HTML scraping — find PDF links inside HTML pages
# ---------------------------------------------------------------------------

def _extract_pdf_links_from_html(html_content: bytes, base_url: str) -> List[str]:
    from urllib.parse import urljoin

    soup     = BeautifulSoup(html_content, "html.parser")
    pdf_urls = []
    seen     = set()

    for tag in soup.find_all(["a", "iframe", "embed", "object"]):
        href = tag.get("href") or tag.get("src") or tag.get("data")
        if not href:
            continue
        if ".pdf" in href.lower() or "pdf" in href.lower():
            absolute = urljoin(base_url, href)
            if absolute not in seen:
                seen.add(absolute)
                pdf_urls.append(absolute)
                logger.info(f"[GEM][HTML] Found PDF link in tag: {absolute}")

    raw_text = html_content.decode("utf-8", errors="ignore")
    pattern  = r'https?://[^\s\'"<>]+\.pdf[^\s\'"<>]*'
    for m in re.findall(pattern, raw_text, re.IGNORECASE):
        if m not in seen:
            seen.add(m)
            pdf_urls.append(m)
            logger.info(f"[GEM][HTML] Found PDF URL via regex: {m}")

    logger.info(f"[GEM][HTML] Total PDF links found: {len(pdf_urls)} | Base: {base_url}")
    return pdf_urls


# ---------------------------------------------------------------------------
# Core PDF downloader + extractor
# ---------------------------------------------------------------------------

def _download_and_extract_pdf(url: str, prefetched_bytes: Optional[bytes] = None) -> str:
    temp_path = None
    try:
        if prefetched_bytes is not None:
            pdf_bytes = prefetched_bytes
        else:
            logger.info(f"[GEM][PDF] Downloading: {url}")
            resp = requests.get(url, timeout=30, verify=False)
            resp.raise_for_status()
            pdf_bytes = resp.content
            logger.info(f"[GEM][PDF] Downloaded {len(pdf_bytes)} bytes | URL: {url}")

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(pdf_bytes)
            temp_path = tmp.name

        full_text  = []
        page_count = 0
        try:
            with fitz.open(temp_path) as doc:
                page_count = doc.page_count
                for page_no in range(page_count):
                    page = doc.load_page(page_no)
                    page_parts = []

                    # Extract tables first — preserves row/column structure
                    try:
                        tabs = page.find_tables()
                        for tab in tabs:
                            df = tab.to_pandas()
                            if not df.empty:
                                page_parts.append(df.to_csv(index=False, sep="|"))
                    except Exception:
                        pass

                    # Extract remaining plain text (non-table regions)
                    plain = page.get_text("text").strip()
                    if plain:
                        page_parts.append(plain)

                    if page_parts:
                        full_text.append("\n".join(page_parts))
        except fitz.FileDataError as e:
            logger.error(f"[GEM][PDF] Invalid PDF: {e} | URL: {url}")
            return f"Error: Invalid PDF — {str(e)}"

        extracted  = "\n".join(full_text).strip()
        char_count = len(extracted)
        logger.info(f"[GEM][PDF] Extracted — Pages: {page_count}, Chars: {char_count} | URL: {url}")

        if char_count == 0 and page_count > 0:
            logger.warning(f"[GEM][PDF] Zero chars from {page_count} pages — trying OCR | URL: {url}")
            ocr_text = _ocr_pdf_bytes(pdf_bytes, url)
            if ocr_text:
                logger.info(f"[GEM][PDF] OCR succeeded — {len(ocr_text)} chars | URL: {url}")
                return ocr_text
            else:
                msg = f"Warning: zero chars even after OCR — scanned PDF or OCR not installed (pages: {page_count})"
                logger.warning(f"[GEM][PDF] {msg} | URL: {url}")
                return msg

        return extracted

    except Exception as e:
        msg = f"Error: {type(e).__name__}: {str(e)}"
        logger.error(f"[GEM][PDF] {msg} | URL: {url}")
        return msg
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)
            logger.info(f"[GEM][PDF] Cleaned up: {temp_path}")


# ---------------------------------------------------------------------------
# Link extractor — handles HTML pages, direct PDFs, scanned PDFs
# ---------------------------------------------------------------------------

def extract_text_from_pdf_link(link: str) -> str:
    logger.info(f"[GEM][LINK] Attempting: {link}")

    try:
        response     = requests.get(link, timeout=30, verify=False)
        http_status  = response.status_code
        content_type = response.headers.get("Content-Type", "unknown")
        content_len  = len(response.content)
        logger.info(
            f"[GEM][LINK] HTTP {http_status} | Content-Type: {content_type} | "
            f"Bytes: {content_len} | URL: {link}"
        )

        response.raise_for_status()

        if "text/html" in content_type.lower():
            logger.info(f"[GEM][LINK] HTML detected — scanning for embedded PDF links | URL: {link}")
            pdf_links = _extract_pdf_links_from_html(response.content, link)

            if pdf_links:
                all_texts = []
                for pdf_url in pdf_links:
                    logger.info(f"[GEM][LINK] Processing embedded PDF: {pdf_url}")
                    result = _download_and_extract_pdf(pdf_url)
                    if result and not result.startswith("Error:") and not result.startswith("Warning:"):
                        all_texts.append(result)
                        logger.info(f"[GEM][LINK] Embedded PDF extracted: {len(result)} chars | {pdf_url}")
                    else:
                        logger.warning(f"[GEM][LINK] Embedded PDF failed: {result} | {pdf_url}")

                if all_texts:
                    combined = "\n\n".join(all_texts)
                    logger.info(f"[GEM][LINK] {len(all_texts)} embedded PDFs extracted, total {len(combined)} chars")
                    return combined
                else:
                    logger.warning(f"[GEM][LINK] PDF links found but all failed to extract | URL: {link}")

            soup      = BeautifulSoup(response.content, "html.parser")
            html_text = soup.get_text(separator="\n", strip=True)
            if html_text and len(html_text) > 100:
                logger.info(f"[GEM][LINK] Extracted {len(html_text)} chars from HTML text | URL: {link}")
                return html_text

            return f"Skipped: HTML page with no useful PDF or text content"

        is_pdf = "pdf" in content_type.lower() or ".pdf" in link.lower()
        if not is_pdf:
            msg = f"Skipped: unrecognized content-type '{content_type}'"
            logger.warning(f"[GEM][LINK] {msg} | URL: {link}")
            return msg

        return _download_and_extract_pdf(link, prefetched_bytes=response.content)

    except requests.exceptions.SSLError as e:
        msg = f"Error: SSL failed — {str(e)}"
        logger.error(f"[GEM][LINK] {msg} | URL: {link}")
        return msg
    except requests.exceptions.ConnectionError as e:
        msg = f"Error: Connection failed — {str(e)}"
        logger.error(f"[GEM][LINK] {msg} | URL: {link}")
        return msg
    except requests.exceptions.Timeout:
        msg = "Error: Timed out after 30s"
        logger.error(f"[GEM][LINK] {msg} | URL: {link}")
        return msg
    except requests.exceptions.HTTPError as e:
        msg = f"Error: HTTP {response.status_code} — {str(e)}"
        logger.error(f"[GEM][LINK] {msg} | URL: {link}")
        return msg
    except Exception as e:
        msg = f"Error: {type(e).__name__}: {str(e)}"
        logger.error(f"[GEM][LINK] {msg} | URL: {link}")
        return msg


# ---------------------------------------------------------------------------
# Main PDF extraction
# ---------------------------------------------------------------------------

def extract_pdf_text_and_links(file_path: str) -> Dict[str, Any]:
    full_text  = []
    all_links  = []
    page_count = 0

    with fitz.open(file_path) as doc:
        page_count = doc.page_count
        for page_no in range(page_count):
            page = doc.load_page(page_no)
            page_parts = []

            try:
                tabs = page.find_tables()
                for tab in tabs:
                    df = tab.to_pandas()
                    if not df.empty:
                        page_parts.append(df.to_csv(index=False, sep="|"))
            except Exception:
                pass

            plain = page.get_text("text").strip()
            if plain:
                page_parts.append(plain)

            if page_parts:
                full_text.append("\n".join(page_parts))
            for link in page.get_links():
                uri = link.get("uri")
                if uri:
                    all_links.append(uri)

    os.makedirs("gem-testing", exist_ok=True)
    if all_links:
        with open("gem-testing/extracted_links.txt", "w") as f:
            for link in all_links:
                f.write(link + "\n")
        logger.info(f"[GEM] Saved {len(all_links)} links to gem-testing/extracted_links.txt")

    extracted_text = "\n".join(full_text)
    logger.info(f"[GEM] Main PDF: {page_count} pages, {len(extracted_text)} chars")
    return {"text": extracted_text, "links": all_links}


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def process_gem_document(file_path: str) -> Dict[str, Any]:
    logger.info(f"[GEM] Starting with file_path: {file_path}")

    try:
        result = extract_pdf_text_and_links(file_path)
        if result is None:
            raise ValueError("extract_pdf_text_and_links returned None")
        text  = result["text"]
        links = result.get("links", [])
    except Exception as e:
        logger.error(f"[GEM] PDF extraction failed: {e}")
        return {"status": "error", "message": f"PDF extraction failed: {str(e)}"}

    logger.info(f"[GEM] Main PDF: {len(text)} chars, {len(links)} links found")

    # ------------------------------------------------------------------
    # Download and extract each linked PDF individually.
    # Each valid (non-empty, non-duplicate) sub-PDF is stored separately
    # so we can feed them to the LLM one-by-one and write a distinct
    # Annexure sheet per PDF.
    # ------------------------------------------------------------------
    sub_pdfs: List[Dict[str, Any]] = []   # [{link, text, index}]
    link_summary: List[Dict]       = []
    duplicate_links_skipped        = 0
    seen_link_content_hashes       = set()

    for idx, link in enumerate(links):
        logger.info(f"[GEM][LINK] Processing link {idx + 1}/{len(links)}: {link}")
        extracted  = extract_text_from_pdf_link(link)
        is_failure = (
            extracted.startswith("Error:")
            or extracted.startswith("Skipped:")
            or extracted.startswith("Warning:")
        )

        if is_failure:
            link_summary.append({"link": link, "status": "failed", "detail": extracted, "chars": 0})
            logger.warning(f"[GEM][LINK] Link {idx + 1} FAILED: {extracted}")
            continue

        char_count   = len(extracted)
        content_hash = hashlib.sha256(extracted.encode("utf-8", errors="ignore")).hexdigest()

        if content_hash in seen_link_content_hashes:
            duplicate_links_skipped += 1
            link_summary.append({"link": link, "status": "duplicate_skipped", "chars": char_count})
            logger.info(f"[GEM][LINK] Link {idx + 1} DUPLICATE CONTENT — skipped ({char_count} chars)")
            continue

        # Zero-char guard (extra safety after extraction)
        if char_count == 0:
            link_summary.append({"link": link, "status": "failed", "detail": "0 chars extracted", "chars": 0})
            logger.warning(f"[GEM][LINK] Link {idx + 1} yielded 0 chars — skipped")
            continue

        seen_link_content_hashes.add(content_hash)
        annexure_index = len(sub_pdfs) + 1
        sub_pdfs.append({"link": link, "text": extracted, "index": annexure_index})
        link_summary.append({"link": link, "status": "success", "chars": char_count, "annexure": annexure_index})
        logger.info(f"[GEM][LINK] Link {idx + 1} SUCCESS — {char_count} chars → Annexure {annexure_index}")

    os.makedirs("gem-testing", exist_ok=True)
    with open("gem-testing/link_summary.json", "w", encoding="utf-8") as f:
        json.dump(link_summary, f, indent=4)
    with open("gem-testing/sub_pdf_texts.txt", "w", encoding="utf-8") as f:
        if not sub_pdfs:
            f.write("No valid sub PDFs extracted.")
        else:
            for sub in sub_pdfs:
                f.write(f"--- SUB PDF {sub['index']} ---\n")
                f.write(f"Link: {sub['link']}\n\n")
                f.write(sub["text"])
                f.write("\n\n")

    successful = sum(1 for s in link_summary if s["status"] == "success")
    failed     = sum(1 for s in link_summary if s["status"] == "failed")
    logger.info(
        f"[GEM][LINK] Summary — Total: {len(links)}, Success: {successful}, "
        f"DuplicateSkipped: {duplicate_links_skipped}, Failed: {failed}"
    )

    # ------------------------------------------------------------------
    # Main PDF: LLM call 1 (Tender Checklist / Asset Details / Resource Cost)
    # ------------------------------------------------------------------
    try:
        cell_map = LLM_processing(text)
    except Exception as e:
        logger.error(f"[GEM] LLM_processing (call1) failed: {e}")
        return {"status": "error", "message": f"Processing failed: {str(e)}"}

    # ------------------------------------------------------------------
    # Sub PDFs: one LLM call per PDF → force-override main sheets +
    # create individual Annexure sheets
    # ------------------------------------------------------------------
    annexure_results: List[Dict[str, Any]] = []   # [{index, link, assets, resources}]

    PRIORITY_SHEETS = ["Asset Details", "Resource Cost"]

    for sub in sub_pdfs:
        ann_idx  = sub["index"]
        ann_link = sub["link"]
        ann_text = sub["text"]

        logger.info(
            f"[GEM][CALL2] Processing Annexure {ann_idx} — "
            f"{len(ann_text)} chars | {ann_link}"
        )

        try:
            # call2 expects a list; wrap single PDF text
            ann_result = call2_LLM_processing([ann_text])
        except Exception as e:
            logger.error(f"[GEM][CALL2] Annexure {ann_idx} failed: {e} | {ann_link}")
            annexure_results.append({
                "index":     ann_idx,
                "link":      ann_link,
                "assets":    [],
                "resources": [],
                "skipped":   True,
                "reason":    str(e),
            })
            continue

        assets    = ann_result.get("Asset Details", [])
        resources = ann_result.get("Resource Cost", [])

        # Nothing extracted → skip entirely, no Annexure sheet
        if not assets and not resources:
            logger.info(
                f"[GEM][CALL2] Annexure {ann_idx} — LLM returned empty extraction, "
                f"no sheet will be created | {ann_link}"
            )
            annexure_results.append({
                "index":     ann_idx,
                "link":      ann_link,
                "assets":    [],
                "resources": [],
                "skipped":   True,
                "reason":    "empty extraction",
            })
            continue

        logger.info(
            f"[GEM][CALL2] Annexure {ann_idx} — "
            f"{len(assets)} assets, {len(resources)} resources | {ann_link}"
        )

        # Force-override the main sheets with the latest sub-PDF data
        for sheet in PRIORITY_SHEETS:
            if sheet in ann_result:
                logger.info(
                    f"[GEM] FORCE OVERRIDE '{sheet}' with Annexure {ann_idx} data"
                )
                cell_map[sheet] = ann_result[sheet]

        annexure_results.append({
            "index":     ann_idx,
            "link":      ann_link,
            "assets":    assets,
            "resources": resources,
            "skipped":   False,
        })

    # Persist annexure extraction summary for debugging
    with open("gem-testing/annexure_results.json", "w", encoding="utf-8") as f:
        json.dump(
            [
                {k: v for k, v in r.items() if k != "assets" and k != "resources"}
                | {"asset_count": len(r.get("assets", [])), "resource_count": len(r.get("resources", []))}
                for r in annexure_results
            ],
            f, indent=4,
        )

    # ------------------------------------------------------------------
    # Normalize sheet names and write Excel
    # ------------------------------------------------------------------
    normalized_cell_map: Dict[str, Any] = {}
    for key, value in cell_map.items():
        normalized_key = key.strip()
        if normalized_key in normalized_cell_map:
            logger.warning(
                f"[GEM] Duplicate sheet key after normalization: '{normalized_key}' — "
                f"keeping later value from '{key}'"
            )
        normalized_cell_map[normalized_key] = value

    try:
        output_path = write_to_excel(normalized_cell_map, annexure_results)
    except Exception as e:
        logger.error(f"[GEM] write_to_excel failed: {e}")
        return {"status": "error", "message": f"Excel write failed: {str(e)}"}

    return {
        "status":       "completed",
        "message":      "PDF processing completed.",
        "output_path":  output_path,
        "link_summary": link_summary,
    }


def gem_processing(file_path: str) -> Dict[str, Any]:
    return process_gem_document(file_path)


def LLM_processing(text: str) -> Dict[str, Any]:
    return process_main_pdf_with_llm(text)


# ---------------------------------------------------------------------------
# LangGraph Chunk Processing
# ---------------------------------------------------------------------------

def process_workflow_chunk(state: ChunkProcessingState) -> ChunkProcessingState:
    if state["current_chunk_index"] >= len(state["all_chunks"]):
        state["is_complete"] = True
        return state

    current_chunk = state["all_chunks"][state["current_chunk_index"]]
    schema        = state["master_schema"]

    logger.info(f"[GEM] Processing chunk {state['current_chunk_index'] + 1}/{len(state['all_chunks'])}")

    previous_context = ""
    if state["accumulated_results"]:
        previous_context = "\n\nPrevious chunk results (for context):\n" + json.dumps(
            state["accumulated_results"], indent=2
        )

    llm = ChatGoogleGenerativeAI(
        model="gemini-2.5-flash",
        google_api_key=os.getenv("GEMINI_API_KEY"),
        temperature=0,
    )

    prompt = ChatPromptTemplate.from_template("""You are an expert tender document analyst filling an Excel workbook from extracted PDF text.

You are processing chunk {chunk_number} of a multi-chunk document.

Return ONLY a valid JSON object. No markdown. No explanation.

Structure:
{{
  "<exact_sheet_name>": {{ "<cell_address>": <value_or_null> }},
  ...
}}

EXCEPTION Asset Details must be a plain JSON array (NOT a dict, NOT wrapped in "rows" or any key):
"Asset Details": [
  {{"asset_type": "Laptop", "brand": null, "qty": 50}},
  {{"asset_type": "Desktop", "brand": "Dell", "qty": 20}}
]

GENERAL RULES
- Use EXACT sheet names and cell addresses from schema.
- Return null if value not found. Do NOT invent values.
- Do NOT duplicate values already present in previous chunks.

TENDER CHECKLIST:
Extract: tender number, bid date, customer/dept name, scope of work, duration, bid due date, EMD/PBG, delivery location, OEM/MAF requirements, compliance.

ASSET DETAILS (plain array):
- Normalize: "Laptops"→"Laptop", "Desktops"→"Desktop", "Printers"→"Printer"
- Total qty rule: If total stated → use total. If only model-wise → sum them.
- ONE entry per asset type. No duplicates.
- Brand: null when using total qty. Set only if one clear brand dominates.
- Types to capture: Desktop, Laptop, Printer, Scanner, Server, UPS, Network Switch, Mouse, Keyboard, Monitor, Webcam, Storage Device

RESOURCE COST:
Extract manpower roles, qualification, qty, monthly rate (only if explicitly stated).

COSTING:
Fill only explicitly stated values. No guessing.

Schema:
{schema}

Chunk {chunk_number}:
{text}

{previous_context}""")

    try:
        chain    = prompt | llm
        response = chain.invoke({
            "schema":           json.dumps(schema, indent=2),
            "text":             current_chunk,
            "chunk_number":     state["current_chunk_index"] + 1,
            "previous_context": previous_context
        })

        raw = response.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()

        chunk_result = json.loads(raw)
        state["chunk_results"].append(chunk_result)

        for sheet_name, cells in chunk_result.items():
            if sheet_name not in state["accumulated_results"]:
                state["accumulated_results"][sheet_name] = [] if isinstance(cells, list) else {}

            if isinstance(cells, list):
                if not isinstance(state["accumulated_results"][sheet_name], list):
                    state["accumulated_results"][sheet_name] = []
                state["accumulated_results"][sheet_name].extend(cells)
            elif isinstance(cells, dict):
                if isinstance(state["accumulated_results"][sheet_name], dict):
                    state["accumulated_results"][sheet_name].update(cells)

        logger.info(f"[GEM] Chunk {state['current_chunk_index'] + 1} processed successfully")

    except json.JSONDecodeError as e:
        logger.error(f"[GEM] JSON parse error chunk {state['current_chunk_index'] + 1}: {e}")
    except Exception as e:
        logger.error(f"[GEM] Error chunk {state['current_chunk_index'] + 1}: {e}")

    state["current_chunk_index"] += 1
    return state


def build_chunk_processing_workflow():
    workflow = StateGraph(ChunkProcessingState)
    workflow.add_node("process_chunk", process_workflow_chunk)
    workflow.set_entry_point("process_chunk")

    def should_continue(state: ChunkProcessingState) -> str:
        if state["is_complete"] or state["current_chunk_index"] >= len(state["all_chunks"]):
            return "end"
        return "process_chunk"

    workflow.add_conditional_edges(
        "process_chunk",
        should_continue,
        {"process_chunk": "process_chunk", "end": "__end__"}
    )

    memory = MemorySaver()
    return workflow.compile(checkpointer=memory)


# ---------------------------------------------------------------------------
# Schema extraction
# ---------------------------------------------------------------------------

def extract_workbook_schema() -> Dict:
    import openpyxl

    file_path = r"templates/template.xlsx"
    wb        = openpyxl.load_workbook(file_path, data_only=False)
    logger.info(f"Workbook sheets: {wb.sheetnames}")

    master_schema = {"workbook_name": "GEM Template", "sheets": {}}

    ws = _find_sheet(wb, "Tender Checklist")
    if ws:
        fields = []
        for row in range(2, 40):
            label = ws[f"A{row}"].value
            if label:
                fields.append({"label": str(label).strip(), "value_cell": f"B{row}", "value": None})
        master_schema["sheets"]["Tender Checklist"] = {
            "sheet_name": ws.title, "type": "key_value_form", "fields": fields
        }
        logger.info(f"[GEM] Schema extracted: '{ws.title}'")
    else:
        logger.warning("[GEM] 'Tender Checklist' sheet not found")

    ws = _find_sheet(wb, "Asset Details")
    if ws:
        master_schema["sheets"]["Asset Details"] = {
            "sheet_name": ws.title, "type": "flat_asset_table",
            "start_row": 3,
            "columns": {"sr_no": "A", "asset_type": "B", "brand": "C", "qty": "D"},
            "rows": []
        }
        logger.info(f"[GEM] Schema extracted: '{ws.title}'")
    else:
        logger.warning("[GEM] 'Asset Details' sheet not found")

    ws = _find_sheet(wb, "Resource Cost")
    if ws:
        resources = []
        for row in range(2, 20):
            sr = ws[f"A{row}"].value
            if sr:
                resources.append({
                    "row": row, "sr_no": sr,
                    "profile_cell": f"B{row}", "qualification_cell": f"C{row}",
                    "qty_cell": f"D{row}", "per_month_cell": f"E{row}",
                    "for_12_month_cell": f"F{row}",
                    "profile": None, "qualification": None,
                    "qty": None, "per_month": None, "for_12_month": None,
                })
        master_schema["sheets"]["Resource Cost"] = {
            "sheet_name": ws.title, "type": "resource_table",
            "resources": resources, "total_cell": "F4"
        }
        logger.info(f"[GEM] Schema extracted: '{ws.title}'")
    else:
        logger.warning("[GEM] 'Resource Cost' sheet not found")

    ws = _find_sheet(wb, "Costing")
    if ws:
        costing_fields = []
        for row in range(1, ws.max_row + 1):
            label = ws[f"C{row}"].value or ws[f"D{row}"].value
            if label:
                costing_fields.append({
                    "label": str(label).strip(), "row": row,
                    "input_cell": f"G{row}", "value": None
                })
        master_schema["sheets"]["Costing"] = {
            "sheet_name": ws.title, "type": "financial_model", "fields": costing_fields
        }
        logger.info(f"[GEM] Schema extracted: '{ws.title}'")
    else:
        logger.warning("[GEM] 'Costing' sheet not found")

    wb.close()

    os.makedirs("gem-testing", exist_ok=True)
    with open("gem-testing/master_schema.json", "w", encoding="utf-8") as f:
        json.dump(master_schema, f, indent=4)

    return master_schema


# ---------------------------------------------------------------------------
# LLM Processing (call 1 — main PDF, full schema)
# ---------------------------------------------------------------------------

def process_main_pdf_with_llm(text: str) -> Dict:
    schema = extract_workbook_schema()
    chunks = split_text_into_chunks(text, chunk_size=180000, overlap=500)

    if not chunks:
        logger.error("[GEM] No chunks created from text")
        return {}

    llm = ChatGoogleGenerativeAI(
        model="gemini-2.5-pro",
        google_api_key=os.getenv("GEMINI_API_KEY"),
        temperature=0,
    )

    initial_state: ChunkProcessingState = {
        "all_chunks":          chunks,
        "current_chunk_index": 0,
        "accumulated_results": {},
        "master_schema":       schema,
        "chunk_results":       [],
        "is_complete":         False,
    }

    graph  = build_chunk_processing_workflow()
    config = {"configurable": {"thread_id": f"gem-session-{uuid4().hex[:8]}"}}

    logger.info(f"[GEM] Starting chunk processing — {len(chunks)} chunks")

    try:
        final_state = None
        for state in graph.stream(initial_state, config):
            if "__end__" in state:
                final_state = state["__end__"]
                break
            if "process_chunk" in state:
                final_state = state["process_chunk"]
        if final_state is None:
            final_state = initial_state
    except Exception as e:
        logger.error(f"[GEM] Workflow failed: {e}")
        final_state = initial_state

    cell_map = final_state.get("accumulated_results", {})

    os.makedirs("gem-testing", exist_ok=True)
    with open("gem-testing/final_cell_map.json", "w", encoding="utf-8") as f:
        json.dump(cell_map, f, indent=4)

    logger.info("[GEM] Chunk processing complete.")
    return cell_map


# ---------------------------------------------------------------------------
# Resource Cost sheet helper functions
# ---------------------------------------------------------------------------

def _find_resource_total_row(ws, start_row: int = 2) -> int:
    for row in range(start_row, ws.max_row + 50):
        e_val = ws[f"E{row}"].value
        f_val = ws[f"F{row}"].value
        if isinstance(e_val, str) and e_val.strip().lower() == "total":
            return row
        if isinstance(f_val, str) and "SUM(" in f_val.upper():
            return row
    return 0


def _copy_row_style(ws, source_row: int, target_row: int, min_col: int = 1, max_col: int = 6):
    for col in range(min_col, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.has_style:
            dst.font       = copy(src.font)
            dst.border     = copy(src.border)
            dst.fill       = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment  = copy(src.alignment)


def _write_resource_cost_sheet(ws, cells: Any) -> int:
    parsed_rows: Dict[int, Dict[str, Any]] = {}

    if isinstance(cells, list):
        for idx, item in enumerate(cells, start=1):
            row = idx + 1
            parsed_rows[row] = {
                "B": item.get("profile"),
                "C": item.get("qualification"),
                "D": item.get("qty"),
                "E": item.get("per_month"),
                "F": item.get("for_12_month"),
            }
    else:
        for cell_addr, value in (cells or {}).items():
            match = re.fullmatch(r"([A-Za-z]+)(\d+)", str(cell_addr).strip())
            if not match:
                continue
            col = match.group(1).upper()
            row = int(match.group(2))
            if row < 2 or col not in {"A", "B", "C", "D", "E", "F"}:
                continue
            parsed_rows.setdefault(row, {})[col] = value

    source_rows = [
        row_num
        for row_num in sorted(parsed_rows.keys())
        if any(parsed_rows[row_num].get(col) is not None for col in ("A", "B", "C", "D", "E", "F"))
    ]
    if not source_rows:
        return 0

    data_start_row = 2
    total_row = _find_resource_total_row(ws, start_row=data_start_row)
    if total_row <= 0:
        total_row = max(ws.max_row + 1, data_start_row + 1)
        ws[f"E{total_row}"] = "Total"

    existing_capacity = max(0, total_row - data_start_row)
    required_rows     = len(source_rows)
    if required_rows > existing_capacity:
        extra_rows = required_rows - existing_capacity
        ws.insert_rows(total_row, amount=extra_rows)
        for new_row in range(total_row, total_row + extra_rows):
            _copy_row_style(ws, source_row=data_start_row, target_row=new_row, min_col=1, max_col=6)
        total_row += extra_rows

    for offset, src_row in enumerate(source_rows):
        target_row = data_start_row + offset
        row_data   = parsed_rows[src_row]
        ws[f"A{target_row}"] = offset + 1
        ws[f"B{target_row}"] = row_data.get("B")
        ws[f"C{target_row}"] = row_data.get("C")
        ws[f"D{target_row}"] = row_data.get("D")
        ws[f"E{target_row}"] = row_data.get("E")
        ws[f"F{target_row}"] = f"=D{target_row}*E{target_row}*12"

    first_data_row = data_start_row
    last_data_row  = data_start_row + required_rows - 1
    ws[f"E{total_row}"] = "Total"
    ws[f"F{total_row}"] = f"=SUM(F{first_data_row}:F{last_data_row})"

    return required_rows * 6


# ---------------------------------------------------------------------------
# Annexure sheet writer
# Writes assets table on top + resources table below in a fresh sheet.
# ---------------------------------------------------------------------------

def _write_annexure_sheet(wb, annexure_index: int, assets: List[dict], resources: List[dict]) -> str:
    """
    Create a new sheet named 'Annexure <N>' in the workbook and populate it
    with the asset and resource data extracted from the corresponding sub-PDF.

    Layout:
      Row 1     : "Asset Details" section header
      Row 2     : Column headers  (Sr No | Asset Type | Brand | Model | Details | Qty | Status | Remarks)
      Row 3+    : Asset rows
      (blank)
      Next row  : "Resource Cost" section header
      Next row+1: Column headers  (Sr No | Profile | Qualification | Qty | Per Month | For 12 Months)
      Next row+2+: Resource rows
      (blank)
      Total row : Total formula for 12-month column

    Returns the created sheet title.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sheet_title = f"Annexure {annexure_index}"
    # Ensure title is unique (openpyxl truncates to 31 chars)
    sheet_title = sheet_title[:31]
    if sheet_title in wb.sheetnames:
        sheet_title = f"Ann{annexure_index}_{uuid4().hex[:4]}"

    ws = wb.create_sheet(title=sheet_title)

    # ---- Style helpers ------------------------------------------------
    header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    section_font   = Font(name="Calibri", bold=True, size=12)
    normal_font    = Font(name="Calibri", size=10)
    center_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    blue_fill      = PatternFill("solid", fgColor="1F4E79")   # asset header
    green_fill     = PatternFill("solid", fgColor="375623")   # resource header
    section_fill   = PatternFill("solid", fgColor="D9E1F2")   # section title row

    thin_side      = Side(style="thin", color="BFBFBF")
    thin_border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def _style_header_cell(cell, fill):
        cell.font      = header_font
        cell.fill      = fill
        cell.alignment = center_align
        cell.border    = thin_border

    def _style_section_cell(cell):
        cell.font      = section_font
        cell.fill      = section_fill
        cell.alignment = left_align

    def _style_data_cell(cell, align=None):
        cell.font      = normal_font
        cell.alignment = align or left_align
        cell.border    = thin_border

    current_row = 1

    # ================================================================
    # SECTION 1: ASSET DETAILS
    # ================================================================
    # Section header spanning all 8 columns
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    section_cell = ws.cell(row=current_row, column=1, value=f"Asset Details — Annexure {annexure_index}")
    _style_section_cell(section_cell)
    current_row += 1

    # Column headers
    asset_headers = ["Sr No", "Asset Type", "Brand", "Model", "Details", "Qty", "Status", "Remarks"]
    for col_idx, header in enumerate(asset_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        _style_header_cell(cell, blue_fill)
    current_row += 1

    # Asset data rows
    if assets:
        for asset in assets:
            ws.cell(row=current_row, column=1, value=asset.get("sr_no"))
            ws.cell(row=current_row, column=2, value=asset.get("asset_type"))
            ws.cell(row=current_row, column=3, value=asset.get("brand"))
            ws.cell(row=current_row, column=4, value=asset.get("model"))
            ws.cell(row=current_row, column=5, value=asset.get("details"))
            ws.cell(row=current_row, column=6, value=asset.get("qty"))
            ws.cell(row=current_row, column=7, value=asset.get("status", ""))
            ws.cell(row=current_row, column=8, value=asset.get("remarks", ""))

            for col_idx in range(1, 9):
                _style_data_cell(
                    ws.cell(row=current_row, column=col_idx),
                    center_align if col_idx in (1, 6) else left_align,
                )
            current_row += 1
    else:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        ws.cell(row=current_row, column=1, value="No asset data extracted from this document.")
        current_row += 1

    current_row += 1   # blank separator row

    # ================================================================
    # SECTION 2: RESOURCE COST
    # ================================================================
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    section_cell = ws.cell(row=current_row, column=1, value=f"Resource Cost — Annexure {annexure_index}")
    _style_section_cell(section_cell)
    current_row += 1

    resource_headers = ["Sr No", "Profile", "Qualification", "Qty", "Per Month (₹)", "For 12 Months (₹)"]
    for col_idx, header in enumerate(resource_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        _style_header_cell(cell, green_fill)
    current_row += 1

    resource_data_start = current_row

    if resources:
        for res in resources:
            ws.cell(row=current_row, column=1, value=res.get("sr_no"))
            ws.cell(row=current_row, column=2, value=res.get("profile"))
            ws.cell(row=current_row, column=3, value=res.get("qualification"))
            ws.cell(row=current_row, column=4, value=res.get("qty"))
            ws.cell(row=current_row, column=5, value=res.get("per_month"))
            ws.cell(row=current_row, column=6, value=f"=D{current_row}*E{current_row}*12"

            )
            for col_idx in range(1, 7):
                _style_data_cell(
                    ws.cell(row=current_row, column=col_idx),
                    center_align if col_idx in (1, 4, 5, 6) else left_align,
                )
            current_row += 1

        # Total row
        resource_data_end = current_row - 1
        ws.cell(row=current_row, column=5, value="Total")
        total_cell = ws.cell(
            row=current_row, column=6,
            value=f"=SUM(F{resource_data_start}:F{resource_data_end})"
        )
        total_cell.font   = Font(name="Calibri", bold=True, size=10)
        total_cell.border = thin_border
        ws.cell(row=current_row, column=5).font   = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=current_row, column=5).border = thin_border
        current_row += 1
    else:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws.cell(row=current_row, column=1, value="No resource cost data extracted from this document.")
        current_row += 1

    # ================================================================
    # Column widths
    # ================================================================
    col_widths = [8, 20, 18, 22, 45, 10, 14, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row height for header rows
    ws.row_dimensions[2].height = 30   # asset header
    # resource header row = row after blank separator
    resource_header_row = resource_data_start - 1
    if resource_header_row > 0:
        ws.row_dimensions[resource_header_row].height = 30

    logger.info(
        f"[GEM] Created sheet '{sheet_title}' — "
        f"{len(assets)} assets, {len(resources)} resources"
    )
    return sheet_title


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_to_excel(cell_map: Dict, annexure_results: Optional[List[Dict[str, Any]]] = None) -> str:
    import shutil
    import openpyxl

    source_file = r"templates/template.xlsx"
    output_dir  = r"templates/outputs"
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"output_{uuid4().hex[:8]}.xlsx")

    shutil.copy(source_file, output_file)
    wb = openpyxl.load_workbook(output_file)

    # ------------------------------------------------------------------
    # Write main template sheets (Tender Checklist, Asset Details, etc.)
    # ------------------------------------------------------------------
    for target_sheet_name, cells in cell_map.items():
        ws = _find_sheet(wb, target_sheet_name)
        if ws is None and target_sheet_name.strip().lower() == "asset details (not merged)":
            template_ws = _find_sheet(wb, "Asset Details")
            if template_ws is not None:
                ws = wb.copy_worksheet(template_ws)
                ws.title = target_sheet_name
        if ws is None:
            logger.warning(f"[GEM] Sheet '{target_sheet_name}' not found — skipping")
            continue

        written = 0
        success = False

        if target_sheet_name.strip().lower() == "asset details":
            if ws[f"F2"].value is None:
                ws[f"F2"] = "Status"
            if ws[f"G2"].value is None:
                ws[f"G2"] = "Remarks"
            logger.info(f"[GEM] Added Status and Remarks headers to '{target_sheet_name}'")

        if isinstance(cells, dict) and set(cells.keys()) == {"rows"} and isinstance(cells.get("rows"), list):
            logger.info(f"[GEM] Unwrapping 'rows' wrapper for '{target_sheet_name}'")
            cells = cells["rows"]

        if (
            isinstance(cells, list)
            and target_sheet_name.strip().lower() == "resource cost"
        ):
            try:
                written = _write_resource_cost_sheet(ws, cells)
                success = True
                logger.info(f"[GEM] Wrote '{target_sheet_name}' - dynamic resource table, {written} cells")
            except Exception as e:
                logger.error(f"[GEM] Resource Cost write failed for '{target_sheet_name}': {e}")

        if not success and isinstance(cells, list):
            try:
                start_row = 3
                for idx, row_data in enumerate(cells, start=1):
                    row_num = start_row + idx - 1
                    details_parts = []
                    model = row_data.get("model")
                    details = row_data.get("details")
                    if model not in (None, ""):
                        details_parts.append(str(model).strip())
                    if details not in (None, ""):
                        details_parts.append(str(details).strip())
                    combined_details = " | ".join(part for part in details_parts if part)
                    ws[f"A{row_num}"] = idx
                    ws[f"B{row_num}"] = row_data.get("asset_type")
                    ws[f"C{row_num}"] = row_data.get("brand")
                    ws[f"D{row_num}"] = combined_details or None
                    ws[f"E{row_num}"] = row_data.get("qty")
                    ws[f"F{row_num}"] = row_data.get("status", "")
                    ws[f"G{row_num}"] = row_data.get("remarks", "")
                written = len(cells) * 7
                success = True
                logger.info(f"[GEM] RAW WRITE '{target_sheet_name}' — {len(cells)} rows written ({written} cells)")
            except Exception as e:
                logger.error(f"[GEM] Raw write failed for '{target_sheet_name}': {e}")

        if (
            not success
            and isinstance(cells, dict)
            and target_sheet_name.strip().lower() == "resource cost"
        ):
            try:
                written = _write_resource_cost_sheet(ws, cells)
                success = True
                logger.info(f"[GEM] Wrote '{target_sheet_name}' - dynamic resource table, {written} cells")
            except Exception as e:
                logger.error(f"[GEM] Resource Cost write failed for '{target_sheet_name}': {e}")

        if not success and isinstance(cells, dict):
            try:
                for cell_addr, value in cells.items():
                    if value is None:
                        continue
                    ws[cell_addr] = value
                    written += 1
                success = True
                logger.info(f"[GEM] Wrote '{target_sheet_name}' — dict, {written} cells")
            except Exception as e:
                logger.error(f"[GEM] Dict write failed for '{target_sheet_name}': {e}")

        if not success and isinstance(cells, list):
            try:
                flat = {}
                for item in cells:
                    if isinstance(item, dict):
                        flat.update(item)
                if not flat:
                    raise ValueError("list payload did not contain any dict values")
                cell_addr_pattern = re.compile(r"^[A-Z]+[0-9]+$", re.IGNORECASE)
                invalid_keys = [key for key in flat.keys() if not cell_addr_pattern.match(str(key))]
                if invalid_keys:
                    raise ValueError(
                        "list payload is row data, not cell map; invalid keys: "
                        + ", ".join(str(key) for key in invalid_keys[:5])
                    )
                for cell_addr, value in flat.items():
                    if value is None:
                        continue
                    ws[cell_addr] = value
                    written += 1
                success = True
                logger.info(f"[GEM] Wrote '{target_sheet_name}' — flattened, {written} cells")
            except Exception as e:
                logger.error(f"[GEM] All write attempts failed for '{target_sheet_name}': {e}")

        if not success:
            logger.error(
                f"[GEM] FAILED '{target_sheet_name}' | type={type(cells)} | "
                f"preview={str(cells)[:300]}"
            )

        logger.info(f"[GEM] Total written to '{ws.title}': {written} cells")

    # ------------------------------------------------------------------
    # Write Annexure sheets (one per valid sub-PDF)
    # ------------------------------------------------------------------
    if annexure_results:
        sheets_created = 0
        for ann in annexure_results:
            if ann.get("skipped"):
                logger.info(
                    f"[GEM] Annexure {ann['index']} skipped — "
                    f"reason: {ann.get('reason', 'unknown')} | {ann.get('link', '')}"
                )
                continue

            assets    = ann.get("assets",    [])
            resources = ann.get("resources", [])

            # Final guard: if both truly empty after all processing, skip sheet
            if not assets and not resources:
                logger.info(
                    f"[GEM] Annexure {ann['index']} — no data to write, sheet not created"
                )
                continue

            sheet_title = _write_annexure_sheet(wb, ann["index"], assets, resources)
            sheets_created += 1
            logger.info(
                f"[GEM] Annexure sheet '{sheet_title}' created — "
                f"{len(assets)} assets, {len(resources)} resources"
            )

        logger.info(f"[GEM] Total Annexure sheets created: {sheets_created}")

    wb.save(output_file)
    wb.close()
    logger.info(f"[GEM] Excel saved → {output_file}")
    return output_file


# ---------------------------------------------------------------------------
# Smart chunker
# ---------------------------------------------------------------------------

from concurrent.futures import ThreadPoolExecutor, as_completed
from rapidfuzz import fuzz

CHUNK_SIZE      = 100_0000000
CHUNK_OVERLAP   = 3_000
FUZZY_THRESHOLD = 85


def _smart_split(text: str, size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    chunks = []
    start  = 0

    while start < len(text):
        end = start + size

        if end >= len(text):
            chunks.append(text[start:])
            break

        search_start = max(start, end - 2000)
        segment      = text[search_start:end]
        blank_pos    = segment.rfind("\n\n")

        if blank_pos != -1:
            cut = search_start + blank_pos + 2
        else:
            newline_pos = segment.rfind("\n")
            cut = (search_start + newline_pos + 1) if newline_pos != -1 else end

        chunks.append(text[start:cut])
        start = max(start + 1, cut - overlap)

    return chunks



# ---------------------------------------------------------------------------
# call2 LLM prompt
# ---------------------------------------------------------------------------

_PROMPT_TEMPLATE = """You are an expert tender document analyst specializing in IT hardware procurement tenders in India (GeM Portal, government PSUs, banks, central government undertakings).

Your task is to extract Asset Details and Resource Cost data from the provided tender document chunk and return it as strict JSON — no markdown, no code fences, no explanation, no extra wrapper keys.

---

## OUTPUT FORMAT (strict JSON, no markdown, no preamble):

{{
  "Asset Details": [
    {{
      "sr_no": <integer, 1-based>,
      "asset_type": <string>,
      "brand": <string or null>,
      "model": <string or null>,
      "details": <string or null>,
      "qty": <integer or null>
    }}
  ],
  "Resource Cost": [
    {{
      "sr_no": <integer, 1-based>,
      "profile": <string>,
      "qualification": <string or null>,
      "qty": <integer or null>,
      "per_month": <number or null>,
      "for_12_month": <number or null>
    }}
  ]
}}

---

## ASSET WHITELIST — extract ONLY these categories:

- Computers / Desktops / All-in-One PCs
- Laptops / Notebooks / Ultrabooks
- Servers (rack, tower, blade)
- Printers (laser, inkjet, dot matrix, thermal, MFP)
- Scanners (flatbed, document, handheld)
- Photocopiers / Reprographic machines
- UPS (Uninterruptible Power Supply)
- Networking: Switches, Routers, Firewalls, Access Points, Load Balancers
- Storage: NAS, SAN, external HDD, tape libraries
- Monitors / Display screens
- Projectors / Interactive flat panels / Smart boards
- KVM switches
- Racks and cabinets
- Thin clients / Zero clients
- Tablets / iPads (only if in IT BOQ)
- Biometric devices / Fingerprint scanners / Attendance systems
- CCTV cameras / IP cameras / DVR / NVR (only if in IT BOQ)
- Video conferencing systems / Webcams / Conference phones
- External storage: pen drives, hard disks (only if in BOQ with qty)
- POS terminals / Barcode scanners (only if in IT BOQ)
- Power Distribution Units (PDU)
- Patch panels / Structured cabling (only if in BOQ with qty)

## STRICTLY FORBIDDEN — never extract:
- Military / defence / weapons / naval / aircraft equipment
- Industrial machinery, vehicles, medical equipment
- Furniture, fixtures, civil/construction materials
- Consumables: toner, ink, paper, cables, cleaning supplies
- Software, licenses, subscriptions, AMC charges
- Manpower / staffing (those go in Resource Cost)

---

## EXHAUSTIVE EXTRACTION — COMPLETENESS RULES:

- MERGED CELL RULE: If a row has qty/date/price columns filled but the asset name/description 
  column is blank or empty, it belongs to the SAME asset as the nearest preceding row that has 
  a name. SUM their quantities into one row. Do NOT create a blank-named row.

You MUST extract EVERY qualifying IT hardware item visible in this chunk. Do not stop early.
This is one segment of a larger document — extract everything present in this segment.

NOTE: This chunk may begin mid-table. If you see rows that clearly belong to a hardware
table even without a visible header, extract them using context clues (columns that look
like brand, model, qty, specs).

---

## MODEL FIELD RULES:

The `model` field must contain the model / part number exactly as it appears in the document,
with ONLY label prefix noise removed.

- Strip ONLY these leading label words (case-insensitive):
  "Model", "Model No", "Model No.", "Model Number",
  "Part No", "Part No.", "Part Number", "SKU", "Code"
  followed by an optional colon or dash.
- Do NOT strip the brand name from the model string — brand is captured separately.
- Do NOT abbreviate, summarise, or invent a model number.
- Examples:
    "Model No. Z4665G"           → model: "Z4665G"
    "Model: HP EliteBook 840 G9" → model: "HP EliteBook 840 G9"
    "SKU: ABC-123-XYZ"           → model: "ABC-123-XYZ"
    Not stated                   → model: null
- If no model / part number is present → model: null

---

## TWO-PASS READING — CRITICAL:

LEVEL 1: Summary BOQ — item name + total qty (e.g. "Laptops — 92")
LEVEL 2: Detail sub-table — individual variants with brand, model, specs, per-variant qty

- Level 2 exists with per-variant qty → one row per variant, use sub-table qty
- Level 2 exists but no per-variant qty → one row per variant, qty = null
- Only Level 1 exists → one row, brand if mentioned, model = null, details = null, qty from Level 1
- NEVER skip sub-tables

---

## QUANTITY RULES:
- Use qty exactly as written
- "as required" / "as per actual" / "to be confirmed" → qty = null
- If no qty is given then default is 1 put 1 DONT MERGE THEM.
- Range given → use lower bound
- Confirmed duplicate after normalization → SUM quantities
---

## BRAND RULES:
- Manufacturer name only (e.g. HP, Dell, Lenovo, Cisco, APC)
- Not stated or unknown → brand = null

---

## DETAILS FIELD RULES:
- Capture ALL useful identifying detail present for the asset.
- Include technical specs when present: CPU, RAM, Storage, Display, OS, speed, interface, print type, scan type, etc.
- If the row only provides a model / part number and no other specs, use that model / part number text in `details` instead of null.
- If the row provides brand + model in one combined string, keep the identifying model text in `details`.
- Prefer fuller identifying text over null whenever any model/spec text is present.
- Format examples:
  "CPU: Intel i5-1235U | RAM: 8GB DDR4 | Storage: 512GB SSD | Display: 15.6 inch | OS: Windows 11"
  "Model: ACER VERITON Z4665G"
  "Model: HP LaserJet Pro M403DN"
- Only include attributes actually present — skip blank columns
- No qty in details
- Use `details = null` only when absolutely no model text and no specs are present

---

## ASSET TYPE RULES:
- Category name only — never include brand/model words
- Wrong: "HP Laptop", "Canon LaserJet" → Right: "Laptop", "Printer"
- Preserve the tender's own terminology exactly if it is a valid category name

---

## RESOURCE COST RULES:

Extract named manpower roles with qty. Look in: Manpower sections, Staffing annexures,
Scope of Work ("bidder shall deploy N engineers"), Pre-Qualification criteria.

- qty = null if not an explicit integer
- qualification = null if not stated
- per_month / for_12_month = null unless explicit INR figure — NEVER calculate
- Nothing found → "Resource Cost": []

---

## ABSOLUTE RULES:
1. Return ONLY the JSON object — no preamble, no markdown, no code fences
2. qty and cost fields = JSON numbers, not strings
3. sr_no = 1-based sequential within this chunk
4. Do NOT fabricate — only extract what is explicitly present
5. Do NOT skip sub-tables
6. Zero assets + zero resources → {{"Asset Details": [], "Resource Cost": []}}

---

DOCUMENT CHUNK:
{text}"""

_PROMPT = ChatPromptTemplate.from_template(_PROMPT_TEMPLATE)


# ---------------------------------------------------------------------------
# call2 LLM Processing (sub-PDFs — Asset Details + Resource Cost)
# ---------------------------------------------------------------------------

def call2_LLM_processing(sub_pdf_texts: List[str]) -> Dict:
    """
    Process a list of sub-PDF texts (treated as a single logical document)
    and return extracted Asset Details and Resource Cost.
    Each call processes ONE sub-PDF's text (callers should pass a single-item list).
    """
    if not sub_pdf_texts:
        logger.info("[GEM] No sub PDF texts for call2")
        return {}

    combined_text = "\n\n--- SUB PDF ---\n\n".join(sub_pdf_texts)
    logger.info(f"[GEM][CALL2] Combined text: {len(combined_text)} chars")

    llm = ChatGoogleGenerativeAI(
        model="gemini-2.5-pro",
        google_api_key=os.getenv("GEMINI_API_KEY"),
        temperature=0,
    )

    chain = _PROMPT | llm

    chunks = _smart_split(combined_text)
    logger.info(f"[GEM][CALL2] Split into {len(chunks)} chunk(s)")

    def process_chunk(idx: int, chunk: str):
        try:
            response = chain.invoke({"text": chunk})
            content  = response.content
            if isinstance(content, list):
                parts = []
                for item in content:
                    if isinstance(item, str):
                        parts.append(item)
                    elif isinstance(item, dict):
                        text = item.get("text")
                        if text:
                            parts.append(str(text))
                    else:
                        text = getattr(item, "text", None)
                        if text:
                            parts.append(str(text))
                raw = "\n".join(part.strip() for part in parts if str(part).strip())
            else:
                raw = str(content or "").strip()

            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
                raw = raw.strip()

            chunk_result = json.loads(raw)
            assets    = chunk_result.get("Asset Details", [])
            resources = chunk_result.get("Resource Cost", [])
            assets    = assets    if isinstance(assets,    list) else []
            resources = resources if isinstance(resources, list) else []

            logger.info(f"[GEM][CALL2] Chunk {idx}: {len(assets)} assets, {len(resources)} resources")
            return idx, assets, resources

        except json.JSONDecodeError as e:
            logger.error(f"[GEM][CALL2] Chunk {idx} JSON parse failed: {e}")
            return idx, [], []
        except Exception as e:
            logger.error(f"[GEM][CALL2] Chunk {idx} LLM call failed: {e}")
            return idx, [], []

    chunk_results: Dict[int, tuple] = {}

    with ThreadPoolExecutor(max_workers=min(len(chunks), 10)) as executor:
        futures = {
            executor.submit(process_chunk, idx, chunk): idx
            for idx, chunk in enumerate(chunks, start=1)
        }
        for future in as_completed(futures):
            idx, assets, resources = future.result()
            chunk_results[idx] = (assets, resources)

    all_assets:    List[dict] = []
    all_resources: List[dict] = []

    for idx in sorted(chunk_results):
        assets, resources = chunk_results[idx]
        all_assets.extend(assets)
        all_resources.extend(resources)
    deduped_assets = all_assets
    deduped_resources = all_resources
    logger.info(
        f"[GEM][CALL2] After dedup — "
        f"assets: {len(deduped_assets)}, resources: {len(deduped_resources)}"
    )
    result = {
        "Asset Details": deduped_assets,
        "Resource Cost": deduped_resources,
    }

    os.makedirs("gem-testing", exist_ok=True)
    with open("gem-testing/call2_raw_result.json", "w", encoding="utf-8") as f:
        json.dump(result, f, indent=4)

    logger.info(f"[GEM][CALL2] Done. Keys: {list(result.keys())}")
    return result
